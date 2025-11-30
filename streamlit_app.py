import streamlit as st
import csv
import io
import random as ra
from typing import List, Tuple, Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string  

# ------------------ CORE LOGIC ------------------ #

def parse_input_text(text: str) -> Tuple[List[str], List[str], List[Dict]]:
    """
    Parse full input text (CSV-like) into:
    - meta_labels: list of B-column strings for first 9 rows
    - meta_values: list of C-column strings for first 9 rows
    - students: list of {sno, regno, name, total}

    Expected input structure in Excel (user view):
    Row1–9:   B: label,  C: value   (A blank)
    Row10+:   A: S.NO, B: REG. NO, C: NAME, D: MARKS

    Corresponding CSV rows:
    Row1–9:   ,SHEET INFO :,CO EVALUATION SHEET
    Row10+:   S.NO,REG. NO,NAME,MARKS
    """
    rows = list(csv.reader(io.StringIO(text)))

    meta_labels: List[str] = []
    meta_values: List[str] = []
    students: List[Dict] = []

    # First 9 rows -> meta (if present) from columns B & C (index 1 & 2)
    for i in range(min(9, len(rows))):
        row = rows[i]
        if not row:
            meta_labels.append("")
            meta_values.append("")
            continue
        label = row[1].strip() if len(row) >= 2 else ""  # B-column
        value = row[2].strip() if len(row) >= 3 else ""  # C-column
        meta_labels.append(label)
        meta_values.append(value)

    # Remaining rows -> student section (including possible header row)
    auto_sno = 1
    for row in rows[9:]:
        if not row or all(not c.strip() for c in row):
            continue
        # need at least 4 columns for student
        if len(row) < 4:
            continue

        # marks in 4th column
        try:
            total = int(row[3].strip())
        except ValueError:
            # header like "MARKS" -> skip
            continue

        # S.NO
        try:
            sno = int(row[0].strip())
        except ValueError:
            sno = auto_sno

        regno = row[1].strip()
        name = row[2].strip()

        students.append(
            {"sno": sno, "regno": regno, "name": name, "total": total}
        )
        auto_sno += 1

    # pad meta to always 9 rows
    while len(meta_labels) < 9:
        meta_labels.append("")
        meta_values.append("")

    return meta_labels, meta_values, students


def assessment_pattern(reg: int, ass: int, dep: Optional[int] = None) -> Tuple[List[int], List[int]]:
    """
    Returns (ms, co) lists for a given regulation and assessment.
    dep: 1 -> S&H, 2 -> Other (only for MODEL).
    NOTE: ass == 6 (Custom) is handled separately via custom_ms/custom_co.
    """
    ms: List[int] = []
    co: List[int] = []

    if reg in (13, 17):
        if ass == 1:
            ms = [2, 2, 2, 2, 2, 16, 16, 8]
            co = [1, 1, 1, 2, 2, 1, 2, 1]
        elif ass == 2:
            ms = [2, 2, 2, 2, 2, 16, 16, 8]
            co = [3, 3, 3, 4, 4, 3, 4, 3]
        elif ass == 3:
            if dep is None:
                raise ValueError("Department is required for MODEL exam.")
            if dep == 1:
                ms = [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 16, 16, 16, 16, 16]
                co = [1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 1, 2, 3, 4, 5]
            elif dep == 2:
                ms = [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 13, 13, 13, 13, 13, 15]
                co = [1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 1, 2, 3, 4, 5, 5]
            else:
                raise ValueError("Invalid department selected.")
        elif ass == 4:
            ms = [20, 20, 20, 20, 20]
            co = [1, 2, 3, 4, 5]
        elif ass == 5:
            ms = [20, 20, 20, 20, 20]
            co = [1, 2, 3, 4, 5]
        else:
            raise ValueError("Invalid assessment number for this regulation.")

    elif reg == 21:
        if ass == 1:
            ms = [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 16, 16, 8]
            co = [1, 1, 1, 1, 2, 2, 2, 2, 3, 3, 1, 2, 3]
        elif ass == 2:
            ms = [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 16, 16, 8]
            co = [4, 4, 4, 4, 5, 5, 5, 5, 3, 3, 4, 5, 3]
        elif ass == 3:
            if dep is None:
                raise ValueError("Department is required for MODEL exam.")
            if dep == 1:
                ms = [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 16, 16, 16, 16, 16]
                co = [1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 1, 2, 3, 4, 5]
            elif dep == 2:
                ms = [2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 13, 13, 13, 13, 13, 15]
                co = [1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 1, 2, 3, 4, 5, 5]
            else:
                raise ValueError("Invalid department selected.")
        elif ass == 4:
            ms = [20, 20, 20, 20, 20]
            co = [1, 2, 3, 4, 5]
        elif ass == 5:
            ms = [20, 20, 20, 20, 20]
            co = [1, 2, 3, 4, 5]
        else:
            raise ValueError("Invalid assessment number for this regulation.")
    else:
        raise ValueError("Unsupported regulation. Use 13, 17, or 21.")

    if len(ms) != len(co):
        raise ValueError("Pattern error: ms and co length mismatch.")

    return ms, co


def _band_bounds(total: int, ms: List[int]) -> Tuple[List[int], List[int]]:
    """
    For a given total and question max marks, compute per-question [lo, hi] ranges
    using the original 'weak/average/good/excellent' logic.
    If impossible, it gracefully falls back to lo=0, hi=ms.
    """
    n = len(ms)
    lo = [0] * n
    hi = [0] * n

    if total == 0:
        return lo, hi

    for i, m in enumerate(ms):
        if total < 41:
            if total > 10:
                lo[i] = 0
                hi[i] = m
            elif total < 6:
                lo[i] = 0
                hi[i] = min(1, m)
            else:  # 6 <= total <= 10
                lo[i] = 0
                hi[i] = min(2, m)
        elif 40 < total < 61:
            if m < 3:
                lo[i] = 0
                hi[i] = m
            else:
                lo[i] = 5
                hi[i] = m
        elif 60 < total < 81:
            if m < 3:
                lo[i] = 1
                hi[i] = m
            else:
                lo[i] = 7
                hi[i] = m
        elif 80 < total < 100:
            if m < 3:
                lo[i] = 2
                hi[i] = m
            else:
                lo[i] = 9
                hi[i] = m
        else:
            lo[i] = m
            hi[i] = m

        lo[i] = max(0, min(lo[i], m))
        hi[i] = max(lo[i], min(hi[i], m))

    sum_lo = sum(lo)
    sum_hi = sum(hi)

    if not (sum_lo <= total <= sum_hi):
        lo = [0] * n
        hi = ms[:]

    return lo, hi


def random_split_total_strict(total: int, ms: List[int]) -> List[int]:
    """
    Strict O(n) random split:
    - sum(vals) == total
    - 0 <= vals[i] <= ms[i]
    - No band logic, only per-question max.
      Ideal for custom patterns like [2,2,2,4,10] for total <= sum(ms).
    """
    if total < 0:
        raise ValueError("Total mark cannot be negative.")

    max_possible = sum(ms)
    if total > max_possible:
        raise ValueError(
            f"Total {total} exceeds maximum possible {max_possible} from split-up."
        )

    n = len(ms)
    vals = [0] * n
    remaining = total

    # suffix_sum[i] = sum of ms from i to end
    suffix_sum = [0] * (n + 1)
    for i in range(n - 1, -1, -1):
        suffix_sum[i] = suffix_sum[i + 1] + ms[i]

    for i in range(n):
        if i == n - 1:
            # last question gets whatever is left
            if remaining > ms[i]:
                raise RuntimeError("Logic error: remaining exceeds question max.")
            vals[i] = remaining
            remaining = 0
        else:
            max_after_i = suffix_sum[i + 1]

            # minimum we MUST give here so that remaining can still be distributed later
            min_x = max(0, remaining - max_after_i)

            # maximum we CAN give here (question limit + remaining)
            max_x = min(ms[i], remaining)

            if min_x > max_x:
                raise RuntimeError(
                    f"No valid allocation at question {i}: min_x={min_x}, max_x={max_x}"
                )

            x = ra.randint(min_x, max_x)
            vals[i] = x
            remaining -= x

    if remaining != 0:
        raise RuntimeError("Random allocation failed to exhaust remaining marks.")

    return vals


def random_split_total(total: int, ms: List[int], use_bands: bool = True) -> List[int]:
    """
    Fast, guaranteed O(n) random split of 'total' into len(ms) parts.

    If use_bands=True:
        - Uses weak/avg/good/excellent band logic via _band_bounds()
        - Good for 100-mark style internal/model exams.

    If use_bands=False:
        - Uses strict 0..ms[i] logic (random_split_total_strict)
        - Good for custom patterns like [2,2,2,4,10].
    """
    if not use_bands:
        return random_split_total_strict(total, ms)

    if total < 0:
        raise ValueError("Total mark cannot be negative.")

    max_possible = sum(ms)
    if total > max_possible:
        raise ValueError(
            f"Total {total} exceeds maximum possible {max_possible} from split-up."
        )

    lo, hi = _band_bounds(total, ms)
    sum_lo = sum(lo)
    sum_hi = sum(hi)

    if not (sum_lo <= total <= sum_hi):
        raise RuntimeError("Internal bounds inconsistency.")

    n = len(ms)
    vals = lo[:]
    remaining = total - sum_lo
    caps = [hi[i] - lo[i] for i in range(n)]

    indices = list(range(n))
    ra.shuffle(indices)

    suffix_caps = [0] * (n + 1)
    for idx in range(n - 1, -1, -1):
        suffix_caps[idx] = suffix_caps[idx + 1] + caps[indices[idx]]

    for pos, i in enumerate(indices):
        if remaining <= 0:
            break

        max_for_this = caps[i]
        max_remaining_for_rest = suffix_caps[pos + 1]
        upper = min(max_for_this, remaining)
        min_extra = max(0, remaining - max_remaining_for_rest)

        if upper < min_extra:
            extra = min_extra
        else:
            extra = ra.randint(min_extra, upper)

        vals[i] += extra
        remaining -= extra

    if remaining != 0:
        raise RuntimeError("Random allocation failed to exhaust remaining marks.")

    return vals


def generate_assessment_excel(
    meta_labels: List[str],
    meta_values: List[str],
    students: List[Dict],
    reg: int,
    ass: int,
    dep: Optional[int],
    ass_name: str,
    ass_short: str,
    custom_ms: Optional[List[int]] = None,
    custom_co: Optional[List[int]] = None,
) -> Tuple[bytes, str]:
    """
    Generate Excel (.xlsx) exactly in the requested layout, with full formatting.
    Adds one TOTAL column after CO columns (header TM, per-student total).
    """
    # Pattern
    if ass == 6:
        if not custom_ms or not custom_co:
            raise ValueError("Custom ms/co not provided for Custom assessment.")
        if len(custom_ms) != len(custom_co):
            raise ValueError("Custom ms and co length mismatch.")
        ms, co = custom_ms, custom_co
    else:
        ms, co = assessment_pattern(reg, ass, dep)

    qno = [i + 1 for i in range(len(ms))]
    q_count = len(ms)

    # CO codes & total max marks per CO
    co_codes = sorted(set(co))
    co_totals = [
        sum(ms[j] for j in range(q_count) if co[j] == c)
        for c in co_codes
    ]

    header_fill = PatternFill(
        start_color="FFD9D9D9",  # light grey
        end_color="FFD9D9D9",
        fill_type="solid"
    )

    subheader_fill = PatternFill(
        start_color="FFFCE4D6",  # light peach for row 13 (optional)
        end_color="FFFCE4D6",
        fill_type="solid"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "CO Evaluation"

    # 1–9: meta header — now in B & C (col 2 & 3), A empty
    for i in range(9):
        row = i + 1
        ws.cell(row=row, column=2, value=meta_labels[i] or "")  # B
        ws.cell(row=row, column=3, value=meta_values[i] or "")  # C

    start_col = 4          # column D for questions
    co_start_col = start_col + q_count
    total_col = co_start_col + len(co_codes)  # one extra TOTAL column
    last_col = total_col

    # ---- Top branding: KGISL INSTITUTE OF TECHNOLOGY ----
    title_fill = PatternFill(
        start_color="FFBDD7EE",  # light blue
        end_color="FFBDD7EE",
        fill_type="solid",
    )

    # Big title on row 5 from D to last_col
    title_cell = ws.cell(row=5, column=start_col, value="KGISL INSTITUTE OF TECHNOLOGY")
    ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=last_col)
    title_cell.font = Font(bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = title_fill

    # subtitle row (row 6)
    subtitle = ws.cell(row=6, column=start_col, value="co-Kreate your Genius")
    ws.merge_cells(start_row=6, start_column=start_col, end_row=6, end_column=last_col)
    subtitle.font = Font(
        bold=False,
        size=12,
    )
    subtitle.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 11: QUESTION / ASSESSMENT NO + CO + TOTAL ---
    ws.cell(row=11, column=3, value="QUESTION / ASSESSMENT NO")
    for j, q in enumerate(qno):
        ws.cell(row=11, column=start_col + j, value=q)

    for k, _ in enumerate(co_codes):
        ws.cell(row=11, column=co_start_col + k, value="CO")

    ws.cell(row=11, column=total_col, value="TOTAL")

    # --- Row 12: COURSE OUTCOME NO + CO numbers + TM in last column ---
    ws.cell(row=12, column=3, value="COURSE OUTCOME NO")
    for j, c in enumerate(co):
        ws.cell(row=12, column=start_col + j, value=c)

    for k, c in enumerate(co_codes):
        ws.cell(row=12, column=co_start_col + k, value=c)

    ws.cell(row=12, column=total_col, value="TM")

    # --- Row 13: S.NO / REG. NO / NAME | MARKS + max marks + total max ---
    ws.cell(row=13, column=1, value="S.NO")
    ws.cell(row=13, column=2, value="REG. NO")
    ws.cell(row=13, column=3, value="NAME | MARKS")

    for j, m in enumerate(ms):
        ws.cell(row=13, column=start_col + j, value=m)

    for k, tot in enumerate(co_totals):
        ws.cell(row=13, column=co_start_col + k, value=tot)

    ws.cell(row=13, column=total_col, value=sum(ms))

    # --- Row 14+: students ---
    start_row = 14
    for idx, stu in enumerate(students):
        r = start_row + idx
        sno = stu["sno"]
        regno = stu["regno"]
        name = stu["name"]
        total = stu["total"]

        ws.cell(row=r, column=1, value=sno)
        ws.cell(row=r, column=2, value=regno)
        ws.cell(row=r, column=3, value=name)

        # Use band logic for normal assessments; strict logic for CUS
        use_bands = (ass != 6)
        spup = random_split_total(total, ms, use_bands=use_bands)

        # Question-wise marks
        for j, mark in enumerate(spup):
            ws.cell(row=r, column=start_col + j, value=mark)

        # CO-wise sums
        for k, c in enumerate(co_codes):
            co_sum = sum(spup[j] for j in range(q_count) if co[j] == c)
            ws.cell(row=r, column=co_start_col + k, value=co_sum)

        # TOTAL column (sum of question marks)
        ws.cell(row=r, column=total_col, value=sum(spup))

    # Verified & Approved By row (after 4 empty rows)
    last_row = start_row + len(students) + 4
    ws.cell(row=last_row, column=3, value="Verified & Approved By")

    # ---------- FORMATTING SECTION ----------

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply border only to non-empty cells
    
    thin = Side(border_style="thin", color="000000")

    border = Border(  # full border
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    left_only_border = Border(  # only left border
        left=thin,
        right=Side(border_style=None),
        top=Side(border_style=None),
        bottom=Side(border_style=None),
    )

    # Apply border with special rules
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=max(last_col, 3)):
        for cell in row:
            if cell.value not in (None, ""):
                col = cell.column  # can be int or letter (merged cells case)

                if isinstance(col, int):
                    col_idx = col
                else:
                    col_idx = column_index_from_string(col)

                row_idx = cell.row

                if col_idx >= 4:  # from column D onwards
                    if 1 <= row_idx <= 9:
                        # D1–D9 → only left border
                        cell.border = left_only_border
                    else:
                        # D10 onwards → normal full border
                        cell.border = border
                else:
                    # Columns A–C → always full border
                    cell.border = border


    
    
    
    # ---- COLOR FILLS ----
    orange_fill = PatternFill(start_color="FFFF9900", end_color="FFFF9900", fill_type="solid")  # Orange
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")  # Yellow

    # ---- B1–B9 = ORANGE, C1–C9 = YELLOW ----
    for r in range(1, 10):
        cell_b = ws.cell(row=r, column=2)  # B column
        cell_c = ws.cell(row=r, column=3)  # C column

        if cell_b.value not in (None, ""):
            cell_b.font = Font(bold=True)
            cell_b.fill = orange_fill

        if cell_c.value not in (None, ""):
            cell_c.font = Font(bold=True)
            cell_c.fill = yellow_fill

    # Rows 10, 11, 12, 13: bold content cells only + row colours
    for r in (10, 11, 12, 13):
        for c in range(1, max(last_col, 3) + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value not in (None, ""):
                cell.font = Font(bold=True)
                cell.alignment = center_align

            # Row-wise background colour
            if r in (10, 11, 12):
                cell.fill = header_fill
            elif r == 13:
                cell.fill = subheader_fill

    # Columns A–C from row 14 to last_row bold + header_fill
    for r in range(14, last_row + 1):
        for c in range(1, 4):  # A, B, C
            cell = ws.cell(row=r, column=c)
            if cell.value not in (None, ""):
                cell.font = Font(bold=True)
                cell.fill = header_fill

    # Verified & Approved By ensure bold
    v_cell = ws.cell(row=last_row, column=3)
    v_cell.font = Font(bold=True)

    # Center align numeric / header area (D onwards)
    for r in range(10, last_row + 1):
        for c in range(start_col, max(last_col, start_col - 1) + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value not in (None, ""):
                cell.alignment = center_align

    # Auto-fit column width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value not in (None, ""):
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2
    # ---- FIXED WIDTH FROM COLUMN D ONWARDS ----
    for col in range(4, last_col + 1):  # 4 = column D
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 6.29

    
    
    # Freeze header (optional, nice structure)
    ws.freeze_panes = "D14"  # keeps header + student name columns visible

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = ass_name + ".xlsx"
    return bio.getvalue(), filename


# ------------------ STREAMLIT UI ------------------ #

st.set_page_config(
    page_title="CO Split-Up Generator",
    layout="centered",
)

st.title("CO Split-Up Generator")
st.caption(
    "Designed by [Sathish Ramanujam](https://professor-sathish.github.io/) · "
    "Powered by IPS Tech Community"
)

st.markdown("---")

# Input mode: Upload or Paste
input_mode = st.radio(
    "Input Mode",
    ["Upload CSV", "Paste table"],
    index=0,
    horizontal=True,
)

uploaded_file = None
pasted_text = ""

col_tmpl, col_input = st.columns([1, 2])

with col_tmpl:
    # NOTE: first 9 rows now have leading comma -> A blank, B label, C value
    template_csv = (
        ",SHEET INFO :,CO EVALUATION SHEET\n"
        ",Course Code :,C302\n"
        ",Course  Name :,PYTHON PROGRAM\n"
        ",Faculty Name :,SATHISH R\n"
        ",Academic Year :,2024-2025 (ODD)\n"
        ",Class :,B.TECH.IT (3RD YEAR 'A')\n"
        ",Regulation :,R2021 - AUC\n"
        ",Total No of Students :,63\n"
        ",ASSESSMENT NAME :,INTERNAL ASSESSMENT-1\n"
        "\n"
        "S.NO,REG. NO,NAME,MARKS\n"
        "1,21CSR001,Student Name,50\n"
    )
    st.download_button(
        label="⬇ Download input template",
        data=template_csv.encode("utf-8"),
        file_name="input_template.csv",
        mime="text/csv",
        help="Template: Row1–9 start at B & C (A blank); then S.NO table.",
    )

with col_input:
    if input_mode == "Upload CSV":
        uploaded_file = st.file_uploader(
            "Upload CSV with header + students",
            type=["csv"],
        )
    else:
        pasted_text = st.text_area(
            "Paste data (header + students)",
            height=260,
            help="Keep row1–9 in B & C (A blank), then S.NO table from A.",
        )

st.markdown("---")

# -------- Parse CSV / pasted text once & show editable header --------

text: Optional[str] = None
meta_labels: List[str] = []
meta_values: List[str] = []
students: List[Dict] = []
edited_meta_values: List[str] = []

if input_mode == "Upload CSV" and uploaded_file is not None:
    try:
        uploaded_file.seek(0)
        text = uploaded_file.read().decode("utf-8")
    except Exception as e:
        st.error(f"Error reading uploaded file: {e}")
elif input_mode == "Paste table" and pasted_text.strip():
    text = pasted_text

if text:
    try:
        meta_labels, meta_values, students = parse_input_text(text)
        st.markdown("### Header details (editable)")
        for i, label in enumerate(meta_labels):
            field_label = label or f"Field {i+1}"
            val = st.text_input(field_label, value=meta_values[i], key=f"meta_{i}")
            edited_meta_values.append(val)
    except Exception as e:
        st.error(f"Error parsing input: {e}")

st.markdown("---")

# Options
col1, col2 = st.columns(2)
with col1:
    reg = st.radio("Regulation", [13, 17, 21], index=2, horizontal=True)
with col2:
    ass_label_to_value = {
        "IA1": 1,
        "IA2": 2,
        "MOD": 3,
        "LAB": 4,
        "PRO": 5,
        "CUS": 6,
    }
    ass_label = st.selectbox(
        "Assessment Type (short name)",
        list(ass_label_to_value.keys()),
        index=0,
    )
    ass = ass_label_to_value[ass_label]

dep = None
custom_ms: Optional[List[int]] = None
custom_co: Optional[List[int]] = None

if ass == 3:
    dep_label = st.radio("Department (for MODEL)", ["S & H", "Other"], horizontal=True)
    dep = 1 if dep_label == "S & H" else 2

# Custom assessment configuration
if ass == 6:
    st.markdown("### Custom Assessment Configuration")
    q_count = st.number_input("Number of Questions", min_value=1, max_value=100, value=5, step=1)
    co_text = st.text_input(
        "CO number for each question (comma-separated)",
        value=",".join(str(i + 1) for i in range(min(q_count, 5))),
        help="Example for 5 questions: 1,2,3,4,5",
    )
    ms_text = st.text_input(
        "Max marks for each question (comma-separated)",
        value=",".join(["2"] * min(q_count, 5)),
        help="Example for 5 questions: 2,2,2,2,2",
    )

    try:
        co_list = [int(x.strip()) for x in co_text.split(",") if x.strip() != ""]
        ms_list = [int(x.strip()) for x in ms_text.split(",") if x.strip() != ""]
        if len(co_list) != q_count or len(ms_list) != q_count:
            st.warning(f"Provide exactly {q_count} CO numbers and {q_count} max marks.")
        else:
            custom_co = co_list
            custom_ms = ms_list
    except ValueError:
        st.warning("CO numbers and max marks must be integers separated by commas.")

ass_name = st.text_input(
    "Output filename (without extension)",
    value="assessment_output",
    max_chars=50,
)

st.markdown("---")

generate_btn = st.button("🚀 Generate CO Split-Up (Excel)", type="primary")

if generate_btn:
    if text is None:
        st.error("Please upload a CSV file or paste data.")
    elif not ass_name.strip():
        st.error("Please enter a valid output filename.")
    elif ass == 6 and (custom_ms is None or custom_co is None):
        st.error("Please provide valid custom CO numbers and max marks.")
    elif not students:
        st.error("No valid student rows found. Check your input format.")
    else:
        try:
            # Use edited values if available, otherwise original
            final_meta_values = edited_meta_values or meta_values

            excel_bytes, filename = generate_assessment_excel(
                meta_labels=meta_labels,
                meta_values=final_meta_values,
                students=students,
                reg=reg,
                ass=ass,
                dep=dep,
                ass_name=ass_name.strip(),
                ass_short=ass_label,
                custom_ms=custom_ms,
                custom_co=custom_co,
            )

            st.success(
                f"Generated CO evaluation sheet for {len(students)} students as '{filename}'."
            )

            st.download_button(
                label="⬇ Download Excel",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            st.error(f"Error: {e}")
